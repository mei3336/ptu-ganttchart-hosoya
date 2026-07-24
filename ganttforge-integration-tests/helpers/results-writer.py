"""
results-writer.py

results/test-results.json（Playwright実行結果。results-tracker.jsが生成）を読み込み、
spec/original-spec.xlsx の「テスト項目一覧」シートに実施結果を書き戻す。

書き込み対象列（既存の列見出しに合わせて自動検出。ハードコードした列番号は使わない）:
  実施日 / 実施者 / 判定 / 証跡 / 不具合票 / 備考

★ 数式は一切上書きしない。「進捗サマリー」シートのOK/NG/未実施集計は
  判定列を参照する既存の数式（COUNTIF等）に任せる。
★ 実行後、必ずExcel/LibreOfficeで開いて数式を再計算させること
  （本スクリプトはopenpyxlのみで完結させるため、キャッシュされた計算値までは更新しない）。

使い方:
    python3 helpers/results-writer.py \
        --results results/test-results.json \
        --spec spec/original-spec.xlsx \
        --out spec/original-spec_実施済み.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--results", default="results/test-results.json")
    parser.add_argument("--spec", default="spec/original-spec.xlsx")
    parser.add_argument("--out", default="spec/original-spec_実施済み.xlsx")
    parser.add_argument("--sheet", default="テスト項目一覧")
    args = parser.parse_args()

    results_path = Path(args.results)
    spec_path = Path(args.spec)

    if not results_path.exists():
        print(f"[ERROR] results file not found: {results_path}", file=sys.stderr)
        sys.exit(1)
    if not spec_path.exists():
        print(f"[ERROR] spec file not found: {spec_path}", file=sys.stderr)
        sys.exit(1)

    results = json.loads(results_path.read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(spec_path)  # formulas preserved (data_only=False)
    ws = wb[args.sheet]

    header_row = 1
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    required = ["ID（大-中-小）", "実施日", "実施者", "判定", "証跡", "不具合票", "備考"]
    missing = [h for h in required if h not in headers]
    if missing:
        print(f"[ERROR] 想定した列見出しが見つかりません: {missing}", file=sys.stderr)
        print(f"        実際の見出し: {list(headers.keys())}", file=sys.stderr)
        sys.exit(1)

    id_col = headers["ID（大-中-小）"]

    updated = 0
    not_found = []

    for row in ws.iter_rows(min_row=header_row + 1):
        test_id = row[id_col - 1].value
        if test_id is None or str(test_id) not in results:
            continue
        r = results[str(test_id)]
        row_idx = row[0].row

        ws.cell(row=row_idx, column=headers["実施日"], value=r.get("実施日", ""))
        ws.cell(row=row_idx, column=headers["実施者"], value=r.get("実施者", ""))
        ws.cell(row=row_idx, column=headers["判定"], value=r.get("判定", ""))
        ws.cell(row=row_idx, column=headers["証跡"], value=r.get("証跡", ""))
        ws.cell(row=row_idx, column=headers["不具合票"], value=r.get("不具合票", ""))
        if r.get("備考"):
            existing = ws.cell(row=row_idx, column=headers["備考"]).value
            note = f"{existing}\n{r['備考']}" if existing else r["備考"]
            ws.cell(row=row_idx, column=headers["備考"], value=note)
        updated += 1

    result_ids = set(results.keys())
    sheet_ids = {
        str(row[id_col - 1].value)
        for row in ws.iter_rows(min_row=header_row + 1)
        if row[id_col - 1].value
    }
    not_found = sorted(result_ids - sheet_ids)

    args_out = Path(args.out)
    args_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args_out)

    print(f"[OK] {updated} 件を書き戻しました → {args_out}")
    if not_found:
        print(f"[WARN] 仕様書に存在しないID（results側のみ）: {not_found}")
    print("[NOTE] 進捗サマリーの自動集計を反映するには、Excel/LibreOfficeで開いて保存し直してください。")


if __name__ == "__main__":
    main()
