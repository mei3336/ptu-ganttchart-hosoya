#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_codemap.py

【目的】
  単一HTML（GanttForge再現アプリの index.html）を静的に解析し、
  「セクション一覧・関数カタログ・ストア相互参照」を Markdown の地図として出力する。

【なぜこのスクリプトなのか（前提→結論）】
  前提：手で書いた目次は、コードが動くたびに古くなる（陳腐化はスナップショット手法の宿命）。
  処理：目次を"手で書く"のではなく"コードから毎回導出する"。
  結論：開発が進んだら再実行するだけで最新の地図が得られ、古くなりようがない。

【使い方】
  python3 generate_codemap.py <入力htmlパス> <出力mdパス>

【設計の約束】
  - コードは一切変更しない（読み取り専用）。
  - 判断を混ぜない。抽出できた事実だけを並べる（"暗黙に補完しない"方針に合わせる）。
"""

import re
import sys
import datetime


# 既知のストア定数 → 実ストア名。DBのSTORE_XXX宣言と対応させる。
STORE_CONST_TO_NAME = {
    "STORE_PROJECTS": "projects",
    "STORE_SCHEDULES": "schedules",
    "STORE_TASKS": "tasks",
    "STORE_COMMENTS": "comments",
    "STORE_CHANGELOG": "changelog",
    "STORE_ISSUES": "issues",
    "STORE_MEMOS": "memos",
    "STORE_QUICKMEMOS": "quickmemos",
    "STORE_SNAPSHOTS": "snapshots",
}

# 本文・コメント中から拾う設計書参照語。
DOC_REFERENCES = [
    "基本設計書",
    "詳細設計書",
    "データモデル設計",
    "シーケンス図集",
    "UIデザイン仕様書",
    "モーダル・ダイアログ一覧",
    "UI差分表",
    "対応表",
]

SECTION_RE = re.compile(r"^\s*//\s*={3,}\s*(.*?)\s*={3,}\s*$")
FUNC_RE = re.compile(r"^\s*(async\s+)?function\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(([^)]*)\)")


def read_lines(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.readlines()


def find_sections(lines):
    """===== タイトル ===== 形式の見出しを、行番号つきで拾う。"""
    sections = []
    for i, line in enumerate(lines):
        m = SECTION_RE.match(line)
        if m:
            sections.append({"line": i + 1, "title": m.group(1).strip()})
    return sections


def section_of(line_no, sections):
    """指定行が属する（直近上方の）セクション見出しを返す。"""
    current = None
    for s in sections:
        if s["line"] <= line_no:
            current = s
        else:
            break
    return current["title"] if current else "(セクション外)"


def leading_comment_block(lines, func_index):
    """
    関数宣言行(func_index)の直上に連続する // コメント行を、上方向に遡って集める。
    空行に当たったら打ち切る。返り値は上から順のコメント文字列リスト。
    """
    block = []
    j = func_index - 1
    while j >= 0:
        stripped = lines[j].strip()
        if stripped.startswith("//"):
            block.append(stripped[2:].strip())
            j -= 1
        elif stripped == "":
            break
        else:
            break
    block.reverse()
    return block


def summarize_purpose(comment_block):
    """
    コメント塊から一行の"目的"を抽出する。優先順位：
      1. 【設計判断：X】 の X（設計意図を最も端的に表す）
      2. 【結果】に続く説明（何を返す/保証するか）
      3. 塊の最初の実質行
    見つからなければ空文字。
    """
    joined = " ".join(comment_block)

    m = re.search(r"【設計判断：(.+?)】", joined)
    if m:
        return m.group(1).strip()

    for c in comment_block:
        m = re.match(r"【結果】(.+)", c)
        if m and m.group(1).strip():
            return m.group(1).strip()

    for c in comment_block:
        if c and not c.startswith("=") and "=====" not in c:
            return c
    return ""


def has_proof_structure(comment_block):
    """前提/処理/結果 の証明構造を持つコメントかどうか。"""
    joined = "".join(comment_block)
    return ("【前提】" in joined) and ("【結果】" in joined or "【処理】" in joined)


STORE_CONST_DECL_RE = re.compile(r"^\s*const\s+STORE_[A-Z]+\s*=")


def function_body_span(lines, func_index, next_func_index):
    """
    関数本文の行範囲を返す。
    宣言行〜次の関数宣言の直前までを基本とするが、その手前に
    「次のセクション見出し」または「次のストア定数宣言(const STORE_XXX =)」が
    現れたらそこで打ち切る。
    これをしないと、あるCRUDブロック末尾の関数が、直後のセクションで宣言される
    次ストアの STORE 定数を誤って"触っている"と拾ってしまう。
    """
    hard_end = next_func_index if next_func_index is not None else len(lines)
    end = hard_end
    for k in range(func_index + 1, hard_end):
        line = lines[k]
        if SECTION_RE.match(line) or STORE_CONST_DECL_RE.match(line):
            end = k
            break
    return "".join(lines[func_index:end])


def stores_touched(body_text):
    """本文が参照しているストアを（宣言順で一意に）返す。"""
    found = []
    for const, name in STORE_CONST_TO_NAME.items():
        if const in body_text and name not in found:
            found.append(name)
    return found


def docs_referenced(body_text, comment_block):
    """本文＋コメントで言及される設計書名を返す。"""
    haystack = body_text + " " + " ".join(comment_block)
    return [d for d in DOC_REFERENCES if d in haystack]


def collect_functions(lines, sections):
    # まず関数宣言の位置を全部拾う。
    func_positions = []
    for i, line in enumerate(lines):
        m = FUNC_RE.match(line)
        if m:
            func_positions.append(
                {
                    "index": i,
                    "line": i + 1,
                    "is_async": bool(m.group(1)),
                    "name": m.group(2),
                    "params": m.group(3).strip(),
                }
            )

    functions = []
    for k, fp in enumerate(func_positions):
        next_index = func_positions[k + 1]["index"] if k + 1 < len(func_positions) else None
        comment = leading_comment_block(lines, fp["index"])
        body = function_body_span(lines, fp["index"], next_index)
        functions.append(
            {
                "line": fp["line"],
                "name": fp["name"],
                "params": fp["params"],
                "is_async": fp["is_async"],
                "section": section_of(fp["line"], sections),
                "purpose": summarize_purpose(comment),
                "proof": has_proof_structure(comment),
                "stores": stores_touched(body),
                "docs": docs_referenced(body, comment),
            }
        )
    return functions


def md_escape(text):
    return text.replace("|", "\\|").replace("\n", " ").strip()


# お手本対応表(GanttForge_お手本実ファイル対応表.xlsx)の「関数対応表」シート。
# 列: 0No 1大分類 2お手本関数 3お手本行数 4… 6実ファイル関数 7実ファイル行数 … 10対応関係 11備考
REF_SHEET_NAME = "関数対応表"


def load_reference_map(xlsx_path):
    """
    お手本対応表を読み、「実ファイル関数名 → お手本の関数/行/対応関係/備考」の辞書を返す。

    【なぜ行番号でなく関数名で結合するのか】
      対応表xlsxの「実ファイル行数」は作成時点の値で、実ファイルを編集すると古くなる。
      一方、関数名は編集後も安定している。だから結合キーは名前にする。
      こうすると実ファイルの行がずれても、お手本との対応は保たれる。

    openpyxl が無い / ファイルが無い / シートが無い場合は None を返し、
    呼び出し側は「お手本対応なし」で通常の地図を出す（動作は止めない）。
    """
    try:
        import openpyxl
    except ImportError:
        print("注意: openpyxl が無いため、お手本対応をスキップします（pip install openpyxl で有効化）。",
              file=sys.stderr)
        return None
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"注意: 対応表を開けませんでした（{e}）。お手本対応をスキップします。", file=sys.stderr)
        return None
    if REF_SHEET_NAME not in wb.sheetnames:
        print(f"注意: 対応表に「{REF_SHEET_NAME}」シートが見つかりません。お手本対応をスキップします。",
              file=sys.stderr)
        return None

    ws = wb[REF_SHEET_NAME]
    ref_map = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        repro = r[6] if len(r) > 6 else None
        if not repro:
            continue
        ref_func = str(r[2] or "").strip()
        ref_lines = str(r[3] or "").strip()
        rel = str(r[10] or "").strip() if len(r) > 10 else ""
        note = str(r[11] or "").strip() if len(r) > 11 else ""
        # 1セルに複数の実ファイル関数名が「、」区切りで入ることがある
        for nm in re.split(r"[、,]", str(repro)):
            nm = nm.strip()
            if nm and nm != "(対応なし)":
                ref_map.setdefault(nm, {
                    "ref_func": ref_func,
                    "ref_lines": ref_lines,
                    "rel": rel,
                    "note": note,
                })
    return ref_map


def build_markdown(src_path, lines, sections, functions, ref_map=None):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    total_lines = len(lines)

    out = []
    out.append(f"# コード地図（自動生成）— {src_path}")
    out.append("")
    out.append(
        f"> このファイルは `generate_codemap.py` が `{src_path}` を解析して自動生成したものです。"
        f"手で編集せず、コードを変えたら再生成してください。"
    )
    out.append("")
    out.append(f"- 生成日時：{now}")
    out.append(f"- 総行数：{total_lines:,} 行")
    out.append(f"- セクション数：{len(sections)}")
    out.append(f"- 関数（function宣言）数：{len(functions)}")
    if ref_map is not None:
        matched_n = sum(1 for f in functions if f["name"] in ref_map)
        out.append(
            f"- お手本対応：`ganttforge.html`（対応表xlsxより、関数名で結合）"
            f"／対応あり {matched_n} 関数"
        )
    out.append("")

    # ---- 1. セクション目次 ----
    out.append("## 1. セクション目次")
    out.append("")
    out.append("各セクションの開始行と、含まれる関数の数。行番号はクリックの目印に。")
    out.append("")
    out.append("| # | 開始行 | セクション | 関数数 |")
    out.append("|---|-------|-----------|-------|")
    func_count_by_section = {}
    for f in functions:
        func_count_by_section[f["section"]] = func_count_by_section.get(f["section"], 0) + 1
    for idx, s in enumerate(sections, 1):
        cnt = func_count_by_section.get(s["title"], 0)
        out.append(f"| {idx} | {s['line']} | {md_escape(s['title'])} | {cnt} |")
    out.append("")

    # ---- 2. 関数カタログ（セクション別） ----
    out.append("## 2. 関数カタログ（セクション別）")
    out.append("")
    out.append(
        "「目的」はコード内コメント（【設計判断】【結果】等）から機械抽出したもの。"
        "「証明構造」は 前提→処理→結果 のコメントが揃っている関数の印。"
    )
    out.append("")
    out.append(
        "> **「直接触るストア」の読み方**："
        "`STORE_xxx` 定数を関数の本文で直接名指ししているストアだけを表示している。"
        "`deleteScheduleCascade` のような複合操作が「—」になるのは、"
        "ストアを直接触らず CRUD 関数を組み合わせて実現しているため（＝組み立て層）。"
        "この「—」は、CRUD層と組み立て層が分離している事実を表す。"
    )
    out.append("")
    if ref_map is not None:
        out.append(
            "> **「お手本(行)」の読み方**："
            "お手本ファイル `ganttforge.html` の対応行（対応表xlsxの値）。"
            "**関数名で結合**しているため、実ファイルの行がずれても対応は保たれる。"
            "「記載なし」＝対応表に個別記載のない関数（多くは再現版で追加したヘルパー）。"
            "対応の“種類”（名称変更/アーキテクチャ変更/分割 など）は §4 に詳しく載せた。"
        )
        out.append("")

    # セクション出現順を保ちつつ関数を束ねる
    ordered_sections = [s["title"] for s in sections]
    # セクション外の関数があれば末尾に足す
    for f in functions:
        if f["section"] not in ordered_sections:
            ordered_sections.append(f["section"])

    seen = set()
    for sec in ordered_sections:
        if sec in seen:
            continue
        seen.add(sec)
        sec_funcs = [f for f in functions if f["section"] == sec]
        if not sec_funcs:
            continue
        out.append(f"### {md_escape(sec)}")
        out.append("")
        if ref_map is not None:
            out.append("| 行 | 関数 | async | 目的 | 直接触るストア | お手本(行) | 参照設計書 | 証明構造 |")
            out.append("|----|------|-------|------|-----------|-----------|-----------|---------|")
        else:
            out.append("| 行 | 関数 | async | 目的 | 直接触るストア | 参照設計書 | 証明構造 |")
            out.append("|----|------|-------|------|-----------|-----------|---------|")
        for f in sec_funcs:
            stores = "・".join(f["stores"]) if f["stores"] else "—"
            docs = "・".join(f["docs"]) if f["docs"] else "—"
            proof = "○" if f["proof"] else ""
            asy = "async" if f["is_async"] else ""
            purpose = md_escape(f["purpose"]) if f["purpose"] else "—"
            if ref_map is not None:
                ent = ref_map.get(f["name"])
                if ent is None:
                    ref_cell = "記載なし"
                elif ent["ref_lines"]:
                    ref_cell = ent["ref_lines"]
                else:
                    # 対応表に名前はあるが、お手本側の行が空（新規/未対応など）
                    ref_cell = "（" + (ent["rel"] or "対応行なし") + "）"
                out.append(
                    f"| {f['line']} | `{f['name']}` | {asy} | {purpose} | {stores} | {md_escape(ref_cell)} | {docs} | {proof} |"
                )
            else:
                out.append(
                    f"| {f['line']} | `{f['name']}` | {asy} | {purpose} | {stores} | {docs} | {proof} |"
                )
        out.append("")

    # ---- 3. ストア別・逆引き ----
    out.append("## 3. ストア別・逆引き（どの関数がどのストアを直接触るか）")
    out.append("")
    out.append(
        "「このストアに関わる処理はどこ？」を引くための逆引き表。"
        "データの流れを追うときの入口になる。"
    )
    out.append("")
    for name in STORE_CONST_TO_NAME.values():
        related = [f for f in functions if name in f["stores"]]
        out.append(f"### `{name}`（{len(related)} 関数）")
        out.append("")
        if related:
            names = ", ".join(f"`{f['name']}`(L{f['line']})" for f in related)
            out.append(names)
        else:
            out.append("—")
        out.append("")

    # ---- 4. お手本対応（詳細・関数名で結合） ----
    if ref_map is not None:
        out.append("## 4. お手本対応（詳細・関数名で結合）")
        out.append("")
        out.append(
            "お手本ファイル `ganttforge.html` との対応。結合キーは**関数名**（実ファイルの行番号は"
            "編集で動くため使わない）。**この再現版は書き直しであり、行の1対1対応ではない**点に注意。"
            "対応の“種類”を「対応関係」列に出す。"
        )
        out.append("")

        # 対応関係の内訳（「そもそも一致は少ない＝書き直し」であることを数字で示す）
        matched = [f for f in functions if f["name"] in ref_map]
        unmatched = [f for f in functions if f["name"] not in ref_map]
        rel_count = {}
        for f in matched:
            rel = ref_map[f["name"]]["rel"] or "(未分類)"
            rel_count[rel] = rel_count.get(rel, 0) + 1
        out.append(
            f"- 現在の関数：{len(functions)} ／ 対応表に名前がある：{len(matched)} ／ "
            f"記載なし（多くは再現版で追加）：{len(unmatched)}"
        )
        out.append("- 対応関係の内訳：")
        for rel, cnt in sorted(rel_count.items(), key=lambda kv: -kv[1]):
            out.append(f"  - {md_escape(rel)}：{cnt}")
        out.append("")

        # 詳細表（セクション順・お手本に対応がある関数のみ）
        out.append("| 現在行 | 実ファイル関数 | お手本 関数 | お手本 行 | 対応関係 | 備考 |")
        out.append("|-------|--------------|------------|----------|---------|------|")
        for f in functions:
            ent = ref_map.get(f["name"])
            if ent is None:
                continue
            note = ent["note"]
            if len(note) > 40:
                note = note[:40] + "…"
            out.append(
                f"| {f['line']} | `{f['name']}` | {md_escape(ent['ref_func'] or '—')} | "
                f"{md_escape(ent['ref_lines'] or '—')} | {md_escape(ent['rel'] or '—')} | {md_escape(note or '—')} |"
            )
        out.append("")

        # 対応表に無い関数（黙って落とさず明示）
        out.append("### 対応表に記載のない関数（再現版で追加した可能性）")
        out.append("")
        if unmatched:
            out.append(", ".join(f"`{f['name']}`(L{f['line']})" for f in unmatched))
        else:
            out.append("なし")
        out.append("")

    return "\n".join(out) + "\n"


def main():
    if len(sys.argv) not in (3, 4):
        print("usage: python3 generate_codemap.py <input.html> <output.md> [お手本対応表.xlsx]",
              file=sys.stderr)
        sys.exit(1)
    src, dst = sys.argv[1], sys.argv[2]
    ref_map = load_reference_map(sys.argv[3]) if len(sys.argv) == 4 else None
    lines = read_lines(src)
    sections = find_sections(lines)
    functions = collect_functions(lines, sections)
    md = build_markdown(src, lines, sections, functions, ref_map)
    with open(dst, "w", encoding="utf-8") as f:
        f.write(md)
    suffix = f" / お手本対応{len(ref_map)}名" if ref_map else ""
    print(f"OK: {dst} を生成（セクション{len(sections)} / 関数{len(functions)}{suffix}）")


if __name__ == "__main__":
    main()
